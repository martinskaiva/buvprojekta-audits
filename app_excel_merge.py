from __future__ import annotations

from io import BytesIO
from typing import Iterable

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CANONICAL_COLUMNS = [
    "Audit_ID", "Document_Filename", "Document_Number", "Page", "Location",
    "Category", "Element_Code", "Comment", "Anchor_Text", "Alternative_Anchor",
    "Reference_Document_Filename", "Reference_Document_Number", "Reference_Page",
    "Reference_Location", "Reference_Evidence_Text", "Annotation_Status",
]

LEGACY_COLUMNS = [
    "note_id", "Nr", "discipline", "target_file", "target_page", "target_area",
    "target_text", "comment_text", "issue_type", "severity", "comparison_files",
    "comparison_pages", "comparison_evidence", "markup_type", "placement_confidence", "status",
]

LEGACY_TO_CANONICAL = {
    "Audit_ID": "note_id",
    "Document_Filename": "target_file",
    "Document_Number": None,
    "Page": "target_page",
    "Location": "target_area",
    "Category": "issue_type",
    "Element_Code": None,
    "Comment": "comment_text",
    "Anchor_Text": "target_text",
    "Alternative_Anchor": None,
    "Reference_Document_Filename": "comparison_files",
    "Reference_Document_Number": None,
    "Reference_Page": "comparison_pages",
    "Reference_Location": None,
    "Reference_Evidence_Text": "comparison_evidence",
    "Annotation_Status": "status",
}

LEGACY_UNUSED_COLUMNS = ["Nr", "discipline", "severity", "markup_type", "placement_confidence"]
KEY_COLUMNS = ["Audit_ID", "Document_Filename", "Comment"]
MIN_SCHEMA_MATCH = 8
PREFERRED_CANONICAL_SHEET = "Audit"

# Audit_ID un Annotation_Status nav satura identitātes daļa.
CONTENT_DUPLICATE_COLUMNS = [
    "Document_Filename", "Document_Number", "Page", "Location", "Category",
    "Element_Code", "Comment", "Anchor_Text", "Alternative_Anchor",
    "Reference_Document_Filename", "Reference_Document_Number", "Reference_Page",
    "Reference_Location", "Reference_Evidence_Text",
]

st.set_page_config(page_title="Audit Excel apvienotājs", page_icon="📊", layout="wide")


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalized_key(value) -> str:
    return " ".join(clean_text(value).lower().split())


def content_fingerprint(row: pd.Series) -> tuple[str, ...]:
    return tuple(normalized_key(row[col]) for col in CONTENT_DUPLICATE_COLUMNS)


def normalize_canonical(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str], list[str]]:
    missing = [c for c in CANONICAL_COLUMNS if c not in df.columns]
    extra = [c for c in df.columns if c not in CANONICAL_COLUMNS]
    work = df.copy()
    for col in missing:
        work[col] = ""
    return work[CANONICAL_COLUMNS].copy(), missing, extra


def convert_legacy(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str], list[str]]:
    missing_source = [c for c in LEGACY_COLUMNS if c not in df.columns]
    extra_source = [c for c in df.columns if c not in LEGACY_COLUMNS]
    converted = pd.DataFrame(index=df.index)
    for canonical_col in CANONICAL_COLUMNS:
        legacy_col = LEGACY_TO_CANONICAL[canonical_col]
        converted[canonical_col] = df[legacy_col] if legacy_col and legacy_col in df.columns else ""
    return converted, missing_source, extra_source


@st.cache_data(show_spinner=False)
def read_workbook(file_bytes: bytes, filename: str) -> tuple[pd.DataFrame, dict]:
    xls = pd.ExcelFile(BytesIO(file_bytes))
    candidates = []

    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name, dtype=object)
        except Exception:
            continue
        df.columns = [str(c).strip() for c in df.columns]
        candidates.append({
            "sheet": sheet_name,
            "df": df,
            "canonical_score": len(set(df.columns) & set(CANONICAL_COLUMNS)),
            "legacy_score": len(set(df.columns) & set(LEGACY_COLUMNS)),
        })

    if not candidates:
        raise ValueError("Excel failā neizdevās nolasīt nevienu šķirkli.")

    preferred = [
        c for c in candidates
        if c["sheet"].strip().lower() == PREFERRED_CANONICAL_SHEET.lower()
        and c["canonical_score"] >= MIN_SCHEMA_MATCH
    ]

    if preferred:
        selected = max(preferred, key=lambda c: c["canonical_score"])
        schema = "GOLD / kanoniskā"
        normalized, missing, extra = normalize_canonical(selected["df"])
        schema_score = selected["canonical_score"]
        source_column_count = len(CANONICAL_COLUMNS)
        unused_legacy = []
    else:
        selected = max(candidates, key=lambda c: max(c["canonical_score"], c["legacy_score"]))
        if selected["canonical_score"] >= selected["legacy_score"] and selected["canonical_score"] >= MIN_SCHEMA_MATCH:
            schema = "GOLD / kanoniskā"
            normalized, missing, extra = normalize_canonical(selected["df"])
            schema_score = selected["canonical_score"]
            source_column_count = len(CANONICAL_COLUMNS)
            unused_legacy = []
        elif selected["legacy_score"] >= MIN_SCHEMA_MATCH:
            schema = "Vecā C2-3 → pārveidota"
            normalized, missing, extra = convert_legacy(selected["df"])
            schema_score = selected["legacy_score"]
            source_column_count = len(LEGACY_COLUMNS)
            unused_legacy = [c for c in LEGACY_UNUSED_COLUMNS if c in selected["df"].columns]
        else:
            raise ValueError(
                "Neizdevās atpazīt ne GOLD, ne veco C2-3 audita struktūru. "
                f"Labākā GOLD sakritība: {selected['canonical_score']}/16; "
                f"labākā C2-3 sakritība: {selected['legacy_score']}/16."
            )

    normalized = normalized.dropna(how="all")
    for col in CANONICAL_COLUMNS:
        normalized[col] = normalized[col].apply(clean_text)

    meaningful_mask = normalized[KEY_COLUMNS].apply(
        lambda row: any(str(v).strip() != "" for v in row), axis=1
    )
    normalized = normalized.loc[meaningful_mask].reset_index(drop=True)
    normalized["_Source_Workbook"] = filename

    ignored_sheets = [name for name in xls.sheet_names if name != selected["sheet"]]
    info = {
        "filename": filename,
        "sheet": selected["sheet"],
        "schema": schema,
        "schema_score": schema_score,
        "source_column_count": source_column_count,
        "rows": len(normalized),
        "missing_columns": missing,
        "extra_columns": extra,
        "unused_legacy_columns": unused_legacy,
        "ignored_sheets": ignored_sheets,
    }
    return normalized, info


def combine_frames(frames: Iterable[pd.DataFrame]) -> pd.DataFrame:
    frames = list(frames)
    if not frames:
        return pd.DataFrame(columns=CANONICAL_COLUMNS + ["_Source_Workbook"])
    return pd.concat(frames, ignore_index=True)


def audit_id_collision_rows(df: pd.DataFrame) -> pd.DataFrame:
    ids = df["Audit_ID"].astype("string").str.strip()
    valid = ids.notna() & (ids != "")
    return df.loc[valid & ids.duplicated(keep=False)].copy()


def add_content_keys(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    work = df.copy()
    keys = []
    for col in CONTENT_DUPLICATE_COLUMNS:
        key_col = f"__key_{col}"
        work[key_col] = work[col].apply(normalized_key)
        keys.append(key_col)
    return work, keys


def content_duplicate_rows(df: pd.DataFrame) -> pd.DataFrame:
    work, keys = add_content_keys(df)
    return df.loc[work.duplicated(subset=keys, keep=False)].copy()


def remove_content_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    work, keys = add_content_keys(df)
    return df.loc[~work.duplicated(subset=keys, keep="first")].reset_index(drop=True)


def repeated_document_groups(df: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    work = df.copy()
    work["__doc_key"] = work["Document_Filename"].apply(normalized_key)
    work = work[work["__doc_key"] != ""]
    result = []
    for doc_key, group in work.groupby("__doc_key", sort=False):
        if group["_Source_Workbook"].nunique() > 1:
            result.append((doc_key, group.copy()))
    return result


def source_comparison(group: pd.DataFrame) -> pd.DataFrame:
    source_sets: dict[str, set[tuple[str, ...]]] = {}
    source_rows: dict[str, int] = {}

    for source, source_group in group.groupby("_Source_Workbook", sort=False):
        source_rows[source] = len(source_group)
        source_sets[source] = {content_fingerprint(row) for _, row in source_group.iterrows()}

    all_sources = list(source_sets)
    rows = []
    for source in all_sources:
        own = source_sets[source]
        other_union: set[tuple[str, ...]] = set()
        for other_source in all_sources:
            if other_source != source:
                other_union |= source_sets[other_source]

        rows.append({
            "Avota Excel": source,
            "Piezīmju rindas": source_rows[source],
            "Unikālas piezīmes avotā": len(own),
            "Sakrīt ar citu avotu": len(own & other_union),
            "Tikai šajā avotā": len(own - other_union),
        })

    return pd.DataFrame(rows)


def apply_document_source_choices(
    df: pd.DataFrame,
    choices: dict[str, str],
) -> pd.DataFrame:
    if not choices:
        return df.copy()

    work = df.copy()
    doc_keys = work["Document_Filename"].apply(normalized_key)
    keep_mask = pd.Series(True, index=work.index)

    for doc_key, choice in choices.items():
        if choice == "Paturēt visus avotus":
            continue
        affected = doc_keys == doc_key
        keep_mask.loc[affected] = work.loc[affected, "_Source_Workbook"] == choice

    return work.loc[keep_mask].reset_index(drop=True)


def assign_global_audit_ids(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy().reset_index(drop=True)
    document_numbers: dict[str, int] = {}
    note_counts: dict[str, int] = {}
    next_document_number = 1
    new_ids = []

    for row_index, row in work.iterrows():
        filename = normalized_key(row["Document_Filename"])
        document_number = normalized_key(row["Document_Number"])
        doc_key = filename or document_number or f"__row_{row_index}"
        if doc_key not in document_numbers:
            document_numbers[doc_key] = next_document_number
            note_counts[doc_key] = 0
            next_document_number += 1
        note_counts[doc_key] += 1
        new_ids.append(f"{document_numbers[doc_key]}.{note_counts[doc_key]}")

    work["Audit_ID"] = new_ids
    return work


def build_excel(df: pd.DataFrame) -> bytes:
    export_df = df[CANONICAL_COLUMNS].copy()
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Audit", index=False)

    output.seek(0)
    wb = load_workbook(output)
    ws = wb["Audit"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 32

    widths = {
        "Audit_ID": 18, "Document_Filename": 52, "Document_Number": 30, "Page": 14,
        "Location": 40, "Category": 24, "Element_Code": 20, "Comment": 68,
        "Anchor_Text": 48, "Alternative_Anchor": 48, "Reference_Document_Filename": 52,
        "Reference_Document_Number": 32, "Reference_Page": 20, "Reference_Location": 40,
        "Reference_Evidence_Text": 68, "Annotation_Status": 22,
    }
    for idx, col_name in enumerate(CANONICAL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col_name, 24)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final.getvalue()


def format_list(values: list[str]) -> str:
    return ", ".join(values) if values else "—"


st.title("📊 Audit Excel apvienotājs")
st.caption("Apvieno GOLD audita Excel un vecos C2-3 tipa audita failus vienā Kywatrace 16 kolonnu Audit failā.")

with st.expander("Kanoniskā 16 kolonnu struktūra", expanded=False):
    st.code(" | ".join(CANONICAL_COLUMNS), language=None)

uploaded_files = st.file_uploader(
    "Augšupielādē vienu vai vairākus Excel failus",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True,
)

if uploaded_files:
    frames, infos, errors = [], [], []

    for uploaded in uploaded_files:
        try:
            frame, info = read_workbook(uploaded.getvalue(), uploaded.name)
            frames.append(frame)
            infos.append(info)
        except Exception as exc:
            errors.append({"Fails": uploaded.name, "Kļūda": str(exc)})

    if infos:
        status_rows = []
        for info in infos:
            status_rows.append({
                "Fails": info["filename"],
                "Izmantotais šķirklis": info["sheet"],
                "Atpazītais formāts": info["schema"],
                "Shēmas sakritība": f'{info["schema_score"]}/{info["source_column_count"]}',
                "Importētās rindas": info["rows"],
                "Trūkstošie avota lauki": format_list(info["missing_columns"]),
                "Neizmantotie vecās shēmas lauki": format_list(info["unused_legacy_columns"]),
                "Ignorētie šķirkļi": format_list(info["ignored_sheets"]),
            })

        st.subheader("Failu pārbaude")
        st.dataframe(pd.DataFrame(status_rows), use_container_width=True, hide_index=True)

        converted_count = sum(info["schema"].startswith("Vecā C2-3") for info in infos)
        if converted_count:
            st.info(f"{converted_count} faili automātiski pārveidoti no vecās C2-3 shēmas uz GOLD struktūru.")

        combined_raw = combine_frames(frames)
        id_collisions = audit_id_collision_rows(combined_raw)
        initial_duplicates = content_duplicate_rows(combined_raw)
        repeated_groups = repeated_document_groups(combined_raw)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Apvienotie faili", len(infos))
        c2.metric("Importētās rindas", len(combined_raw))
        c3.metric("Audit_ID sadursmju rindas", len(id_collisions))
        c4.metric("Satura dublikātu rindas", len(initial_duplicates))

        st.caption(
            "Audit_ID sadursme nozīmē tikai to, ka dažādos avota Excel izmantots vienāds ID; "
            "tā pati par sevi nav satura dublikāts."
        )

        source_choices: dict[str, str] = {}
        if repeated_groups:
            st.warning(
                f"Atrasti {len(repeated_groups)} PDF dokumenti, kuru piezīmes ir vairākos augšupielādētajos Excel failos. "
                "Katram dokumentam zemāk vari izvēlēties, kuru avotu izmantot."
            )
            st.subheader("Atkārtoto dokumentu avotu salīdzinājums")

            for index, (doc_key, group) in enumerate(repeated_groups, start=1):
                document_name = group.iloc[0]["Document_Filename"]
                sources = list(dict.fromkeys(group["_Source_Workbook"].tolist()))
                comparison = source_comparison(group)

                with st.expander(f"{index}. {document_name} — {len(sources)} Excel avoti", expanded=False):
                    st.dataframe(comparison, use_container_width=True, hide_index=True)
                    choice = st.selectbox(
                        "Ko izmantot gala failā?",
                        options=["Paturēt visus avotus"] + sources,
                        index=0,
                        key=f"source_choice_{index}_{doc_key[:40]}",
                    )
                    source_choices[doc_key] = choice

                    selected_preview = group if choice == "Paturēt visus avotus" else group[group["_Source_Workbook"] == choice]
                    st.caption(f"Ar šo izvēli no šī PDF gala failā paliks {len(selected_preview)} rindas.")

        selected_df = apply_document_source_choices(combined_raw, source_choices)
        after_selection_duplicates = content_duplicate_rows(selected_df)
        removed_by_source_selection = len(combined_raw) - len(selected_df)

        if repeated_groups:
            s1, s2, s3 = st.columns(3)
            s1.metric("Rindas pēc avotu izvēles", len(selected_df))
            s2.metric("Atmestas ar avotu izvēli", removed_by_source_selection)
            s3.metric("Atlikušo satura dublikātu rindas", len(after_selection_duplicates))

        if not after_selection_duplicates.empty:
            duplicate_groups = len(after_selection_duplicates.drop_duplicates(subset=CONTENT_DUPLICATE_COLUMNS))
            st.warning(
                f"Pēc avotu izvēles vēl ir {len(after_selection_duplicates)} rindas, kas ietilpst "
                f"{duplicate_groups} identisku piezīmju grupās."
            )
            with st.expander("Parādīt atlikušos satura dublikātus"):
                display_cols = ["_Source_Workbook"] + CANONICAL_COLUMNS
                st.dataframe(after_selection_duplicates[display_cols], use_container_width=True, hide_index=True)

        remove_duplicates = st.checkbox(
            "Gala failā noņemt identiskas satura piezīmes (paturēt pirmo)",
            value=False,
            help="Audit_ID un Annotation_Status netiek izmantoti satura dublikāta noteikšanai.",
        )

        final_df = remove_content_duplicates(selected_df) if remove_duplicates else selected_df.copy()
        removed_duplicates = len(selected_df) - len(final_df)

        renumber_ids = st.checkbox(
            "Gala failam izveidot jaunus unikālus Audit_ID pēc dokumentiem",
            value=True,
            help="Pirmajam dokumentam 1.1, 1.2..., nākamajam 2.1, 2.2... utt.",
        )
        if renumber_ids:
            final_df = assign_global_audit_ids(final_df)

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Gala rindas", len(final_df))
        m2.metric("Atmestas ar avotu izvēli", removed_by_source_selection)
        m3.metric("Noņemtie identiskie dublikāti", removed_duplicates)
        m4.metric("PDF dokumenti", final_df["Document_Filename"].replace("", pd.NA).nunique())

        st.subheader("Priekšskatījums")
        st.dataframe(final_df[CANONICAL_COLUMNS].head(100), use_container_width=True, hide_index=True)

        excel_bytes = build_excel(final_df)
        st.download_button(
            "⬇️ Lejupielādēt apvienoto Excel",
            data=excel_bytes,
            file_name="combined_audit_16_columns.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

    if errors:
        st.error("Daļu failu neizdevās importēt.")
        st.dataframe(pd.DataFrame(errors), use_container_width=True, hide_index=True)
else:
    st.info("Vari vienlaikus likt gan GOLD failus, gan vecos C2-3 audita failus.")
